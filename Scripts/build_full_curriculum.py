"""
Build full G6–G9 chapter-aligned skill hierarchy + keyword catalogs.

Sources of truth (Sri Lankan Educational Publications Department textbooks):
  Grade 6 single volume ch.1–11
  Grade 7 Part I ch.1–10 + Part II ch.11–19
  Grade 8 Part I ch.1–8 + Part II ch.9–15
  Grade 9 Part I ch.1–9 + Part II ch.10–19

Writes:
  - Data/Skill-Heirarchies-G6-G9.xlsx  (runtime source)
  - Data/Skill-Heirarchies-G6-G9-Full-Chapters.xlsx  (shareable team copy)
  - Data/chapter_ids_g6_g9.csv  (shared G{grade}_C{chapter} keys for Component 2)
  - FastAPI-Backend/curriculum_topics.py  (keywords / boosts for RAG + routing)
"""

from __future__ import annotations

import csv
import json
import re
from pathlib import Path
from typing import Any

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font
except ImportError as exc:  # pragma: no cover
    raise SystemExit("openpyxl required: pip install openpyxl") from exc


PROJECT_ROOT = Path(__file__).resolve().parent.parent
DATA_DIR = PROJECT_ROOT / "Data"
BACKEND_DIR = PROJECT_ROOT / "FastAPI-Backend"


def _kw(*parts: str) -> list[str]:
    seen: list[str] = []
    for p in parts:
        t = p.strip().lower()
        if t and t not in seen:
            seen.append(t)
    return seen


# Each entry:
# (grade, chapter, chapter_title, part, domain, concept, skill_label, keywords[])
# topic_id = G{g}_C{ch}_{domain}_{concept}
#
# At least one skill per textbook chapter; usually two learning-focus skills.
CURRICULUM: list[tuple[Any, ...]] = [
    # ─── GRADE 6 (11 chapters) ─────────────────────────────────────────────
    (6, 1, "Wonders of the Living World", "full", "ORG", "CHARS",
     "Characteristics of organisms",
     _kw("characteristics of organisms", "living", "growth", "nutrition", "reproduction",
         "respiration", "excretion", "sensitivity", "movement", "alive")),
    (6, 1, "Wonders of the Living World", "full", "ORG", "DIFF",
     "Differences between plants and animals",
     _kw("plants and animals", "plant", "animal", "differences", "photosynthesis",
         "locomotion", "chlorophyll", "autotroph")),
    (6, 2, "Things Around Us", "full", "MAT", "STATES",
     "States of matter",
     _kw("states of matter", "solid", "liquid", "gas", "particle", "melting",
         "freezing", "evaporation", "condensation")),
    (6, 2, "Things Around Us", "full", "MAT", "PROPS",
     "Specific properties of solid matter",
     _kw("properties of solid", "hardness", "malleability", "ductility", "density",
         "transparency", "absorbency", "conductivity")),
    (6, 3, "Water as a Natural Resource", "full", "WAT", "STATES",
     "States and types of water",
     _kw("water", "states of water", "fresh water", "salt water", "salinity",
         "availability of water", "ice", "steam", "vapour")),
    (6, 3, "Water as a Natural Resource", "full", "WAT", "IMPORT",
     "Importance of water and conservation",
     _kw("importance of water", "conserve water", "limited resource", "water cycle",
         "drinking water", "drought")),
    (6, 4, "Energy in Day to Day Life", "full", "ENE", "SOURCES",
     "Energy sources and their applications",
     _kw("energy sources", "energy", "fuel", "solar", "wind", "electric",
         "applications of energy", "day to day energy")),
    (6, 4, "Energy in Day to Day Life", "full", "ENE", "FORMS",
     "Forms of energy in daily life",
     _kw("forms of energy", "heat energy", "light energy", "sound energy",
         "mechanical energy", "chemical energy")),
    (6, 5, "Light and Vision", "full", "LIG", "SEE",
     "How we see; sources and transmission of light",
     _kw("light", "vision", "how can we see", "source of light", "transparent",
         "opaque", "translucent", "transmission of light")),
    (6, 5, "Light and Vision", "full", "LIG", "RAYS",
     "Light rays, beams, and applications",
     _kw("light ray", "light beam", "reflection everyday", "applications of light",
         "shadow", "mirror")),
    (6, 6, "Sound and Hearing", "full", "SOU", "PRODUCE",
     "Producing and hearing sounds",
     _kw("sound", "hearing", "producing sound", "vibration", "ear", "listen")),
    (6, 6, "Sound and Hearing", "full", "SOU", "DIVERSE",
     "Diversity of sounds; music and noise",
     _kw("music", "noise", "diversity of sounds", "pitch", "loudness",
         "equipment to produce sound")),
    (6, 7, "Magnets", "full", "MAG", "POLES",
     "Magnetic poles, types, and behaviour",
     _kw("magnet", "magnetic pole", "north pole", "south pole", "types of magnets",
         "behaviour of a magnet", "bar magnet")),
    (6, 7, "Magnets", "full", "MAG", "FORCES",
     "Magnetic interactions and forces",
     _kw("magnetic force", "attraction", "repulsion", "like poles",
         "unlike poles", "magnetic field basic")),
    (6, 8, "Electricity for a Comfortable Life", "full", "ELE", "CIRCUITS",
     "Preparation of circuits and generating electricity",
     _kw("electric circuit", "circuit", "cell", "battery", "bulb", "switch",
         "generating electricity", "wire", "series", "parallel")),
    (6, 8, "Electricity for a Comfortable Life", "full", "ELE", "CONDINS",
     "Conductors and insulators; safety",
     _kw("conductor", "insulator", "conducting", "insulating", "metal", "plastic",
         "rubber", "conservation of electricity", "electrical safety")),
    (6, 9, "Heat and Its Effects", "full", "HEA", "GEN",
     "Heat generation and effects",
     _kw("heat", "heat generation", "temperature basic", "effects of heat",
         "hot", "cold", "expand heat")),
    (6, 9, "Heat and Its Effects", "full", "HEA", "ENV",
     "Effects of heat on the environment",
     _kw("heat environment", "global warming basic", "heat pollution",
         "effects of heat to the environment")),
    (6, 10, "Food Related Interactions", "full", "FOO", "INTER",
     "Food related interactions in ecosystems",
     _kw("food", "food chain", "food web", "producer", "consumer",
         "decomposer", "interaction food")),
    (6, 10, "Food Related Interactions", "full", "FOO", "NUTR",
     "Nutrition relationships among organisms",
     _kw("nutrition", "herbivore", "carnivore", "omnivore", "feeding",
         "energy flow food")),
    (6, 11, "Weather and Climate", "full", "WEA", "WEATHER",
     "Weather, climate, and measurement",
     _kw("weather", "climate", "rainfall", "temperature weather", "humidity",
         "weather apparatus", "rain gauge")),
    (6, 11, "Weather and Climate", "full", "WEA", "DISASTER",
     "Natural disasters from climatic changes",
     _kw("natural disaster", "flood", "drought", "climatic change",
         "storm", "weather disaster")),

    # ─── GRADE 7 Part I (1–10) ─────────────────────────────────────────────
    (7, 1, "Plant Diversity", "I", "PLA", "DIVER",
     "Morphological features and plant diversity",
     _kw("plant diversity", "morphological", "flowering plant", "root", "stem",
         "leaf", "flower parts", "diversity of plant parts")),
    (7, 1, "Plant Diversity", "I", "PLA", "CLASSIF",
     "Monocotyledonous and dicotyledonous plants",
     _kw("monocot", "dicot", "monocotyledonous", "dicotyledonous", "cotyledon",
         "plant classification")),
    (7, 2, "Static Electricity", "I", "STA", "CHARGES",
     "Charging objects and types of static charge",
     _kw("static electricity", "charging", "static charge", "positive charge",
         "negative charge", "friction charging", "electrostatic")),
    (7, 2, "Static Electricity", "I", "STA", "CAPACIT",
     "Capacitors and static electricity phenomena",
     _kw("capacitor", "capacitance", "store charge", "static phenomena",
         "electroscope")),
    (7, 3, "Generation of Electricity", "I", "ELE", "SOURCES",
     "Sources of electricity generation",
     _kw("generation of electricity", "sources of electricity", "hydro power",
         "thermal power", "solar electricity", "wind turbine", "generator")),
    (7, 3, "Generation of Electricity", "I", "ELE", "CURRENTS",
     "Direct current and alternating current",
     _kw("direct current", "alternating current", "dc", "ac", "electric current",
         "ammeter")),
    (7, 4, "Functions of Water", "I", "WAT", "SOLVENT",
     "Water as a solvent and medium of life",
     _kw("water as solvent", "universal solvent", "dissolve", "solute",
         "solution", "medium of life", "water functions")),
    (7, 4, "Functions of Water", "I", "WAT", "COOLANT",
     "Water as a coolant",
     _kw("coolant", "cooling", "heat capacity of water", "thermal properties of water")),
    (7, 5, "Acids and Bases", "I", "ACI", "IDENTIF",
     "Identification of acids and bases",
     _kw("acid", "base", "alkali", "identification of acids", "litmus",
         "laboratory acid", "home acid")),
    (7, 5, "Acids and Bases", "I", "ACI", "INDICAT",
     "Indicators and neutralization ideas",
     _kw("indicator", "ph", "neutralization", "universal indicator",
         "acid base reaction")),
    (7, 6, "Animal Diversity", "I", "ANI", "CLASSIF",
     "Vertebrates, invertebrates, dichotomous keys",
     _kw("vertebrate", "invertebrate", "animal classification", "dichotomous key",
         "animal diversity")),
    (7, 6, "Animal Diversity", "I", "ANI", "ADAPTAT",
     "Animal adaptations to environment",
     _kw("adaptation", "animal adaptation", "habitat", "camouflage",
         "survive environment")),
    (7, 7, "Forms of Energy and Uses", "I", "ENE", "FORMS",
     "Forms of energy (kinetic, potential, thermal, …)",
     _kw("forms of energy", "kinetic energy", "potential energy", "electrical energy",
         "sound energy", "light energy", "thermal energy", "chemical energy")),
    (7, 7, "Forms of Energy and Uses", "I", "ENE", "TRANSF",
     "Energy transformation and uses",
     _kw("energy transformation", "energy conversion", "uses of energy",
         "energy transfer")),
    (7, 8, "The Nature of the Earth", "I", "EAR", "STRUCT",
     "Structure of the Earth",
     _kw("structure of the earth", "crust", "mantle", "core", "earth layers",
         "internal structure")),
    (7, 8, "The Nature of the Earth", "I", "EAR", "TECTON",
     "Tectonic plates and plate tectonics",
     _kw("tectonic", "plate tectonics", "plate movement", "earthquake basic",
         "continental plate")),
    (7, 9, "Light", "I", "LIG", "SHADOWS",
     "Formation of umbra and penumbra",
     _kw("umbra", "penumbra", "shadow", "formation of shadows", "light shadow")),
    (7, 9, "Light", "I", "LIG", "MIRRORS",
     "Images in plane and curved mirrors",
     _kw("plane mirror", "curved mirror", "reflection", "image formation",
         "concave mirror", "convex mirror")),
    (7, 10, "The Correct Use of the Microscope", "I", "MIC", "LIGHT",
     "Simple and compound light microscopes",
     _kw("microscope", "compound microscope", "simple microscope", "magnification",
         "resolving power", "objective lens")),
    (7, 10, "The Correct Use of the Microscope", "I", "MIC", "ELECTR",
     "Electron microscope characteristics",
     _kw("electron microscope", "resolution electron", "sem", "tem",
         "high resolution microscope")),

    # ─── GRADE 7 Part II (11–19) ───────────────────────────────────────────
    (7, 11, "Sound", "II", "SOU", "PRODUCE",
     "Production of sound",
     _kw("production of sound", "vibration sound", "sound source", "how sound is produced")),
    (7, 11, "Sound", "II", "SOU", "PROPAG",
     "Propagation of sound",
     _kw("propagation of sound", "sound medium", "sound wave travel", "echo basic",
         "sound through air")),
    (7, 12, "Biological Processes", "II", "BIO", "LEVELS",
     "Organisational levels of life",
     _kw("organisational levels", "cell", "tissue", "organ", "system",
         "levels of organisation")),
    (7, 12, "Biological Processes", "II", "BIO", "SYSTEMS",
     "Systems of the human body",
     _kw("human body systems", "digestive system intro", "respiratory system intro",
         "circulatory intro", "biological processes body")),
    (7, 13, "Atmosphere", "II", "ATM", "LAYERS",
     "Layers of the atmosphere",
     _kw("atmosphere", "troposphere", "stratosphere", "layers of atmosphere",
         "atmospheric layers")),
    (7, 13, "Atmosphere", "II", "ATM", "AIR",
     "Air and its components",
     _kw("air", "oxygen", "nitrogen", "carbon dioxide air", "composition of air",
         "components of air")),
    (7, 14, "Heat and Temperature", "II", "HEA", "MEASURE",
     "Measuring temperature and thermometers",
     _kw("temperature", "thermometer", "measuring temperature", "celsius",
         "heat and temperature")),
    (7, 14, "Heat and Temperature", "II", "HEA", "TRANSF",
     "Heat transfer and convection applications",
     _kw("heat transfer", "conduction", "convection", "radiation heat",
         "convectional currents")),
    (7, 15, "Soil", "II", "SOI", "TYPES",
     "Types and composition of soil",
     _kw("soil", "types of soil", "sand", "clay", "loam", "composition of soil",
         "soil particles")),
    (7, 15, "Soil", "II", "SOI", "EROSION",
     "Soil erosion",
     _kw("soil erosion", "erode soil", "conservation of soil", "wash away soil")),
    (7, 16, "Force and Motion", "II", "FOR", "DIST",
     "Distance and displacement",
     _kw("distance", "displacement", "motion", "position", "path length")),
    (7, 16, "Force and Motion", "II", "FOR", "FORCE",
     "Force basics",
     _kw("force", "push", "pull", "newton basic", "effect of force", "force and motion")),
    (7, 17, "Nutrients in Food", "II", "NUT", "FOOD",
     "Food and nutrients",
     _kw("nutrients", "carbohydrate", "protein", "fat", "vitamin", "mineral food",
         "balanced diet")),
    (7, 17, "Nutrients in Food", "II", "NUT", "TESTS",
     "Food tests to identify nutrients",
     _kw("food test", "iodine test", "biuret", "sudan", "identify nutrients",
         "food sample test")),
    (7, 18, "Minerals and Rocks", "II", "ROC", "KINDS",
     "Features and kinds of rocks and minerals",
     _kw("rock", "mineral", "igneous", "sedimentary", "metamorphic",
         "features of minerals")),
    (7, 18, "Minerals and Rocks", "II", "ROC", "CYCLE",
     "Rock weathering and rock cycle",
     _kw("rock cycle", "weathering", "rock weathering", "erosion rocks")),
    (7, 19, "Sources of Energy", "II", "ENE", "RENEW",
     "Renewable and non-renewable energy sources",
     _kw("renewable", "non-renewable", "fossil fuel", "solar energy", "wind energy",
         "sources of energy", "sustainable energy")),
    (7, 19, "Sources of Energy", "II", "ENE", "SUSTAIN",
     "Sustainable usage of energy sources",
     _kw("sustainable energy", "energy conservation", "efficient use of energy",
         "save energy")),

    # ─── GRADE 8 Part I (1–8) ──────────────────────────────────────────────
    (8, 1, "Importance of Microorganisms", "I", "MIC", "INTRO",
     "Microorganisms and their diversity",
     _kw("microorganism", "bacteria", "yeast", "fungi", "protozoa", "algae micro",
         "microscopic organisms")),
    (8, 1, "Importance of Microorganisms", "I", "MIC", "EFFECTS",
     "Effects of microorganisms on food and humans",
     _kw("spoilage", "food spoilage", "microbes on food", "pathogen",
         "effects of microorganisms", "useful microorganisms")),
    (8, 2, "Animal Classification", "I", "ANI", "INVERT",
     "Main invertebrate groups",
     _kw("invertebrate", "arthropod", "mollusc", "annelid", "cnidarian",
         "invertebrate groups")),
    (8, 2, "Animal Classification", "I", "ANI", "VERT",
     "Main vertebrate groups",
     _kw("vertebrate", "fish", "amphibian", "reptile", "bird", "mammal",
         "vertebrate groups")),
    (8, 3, "Diversity and Functions of Plant Parts", "I", "PLA", "LEAVES",
     "Diversity and functions of plant leaves",
     _kw("leaf", "leaves", "functions of leaves", "photosynthesis leaf",
         "leaf diversity")),
    (8, 3, "Diversity and Functions of Plant Parts", "I", "PLA", "STEMROOT",
     "Diversity and functions of stems and roots",
     _kw("stem", "root", "functions of stem", "functions of root",
         "plant parts", "storage root")),
    (8, 4, "Properties of Matter", "I", "MAT", "PARTICLE",
     "Discontinuous nature of matter",
     _kw("particle nature", "discontinuous nature", "atoms and molecules intro",
         "matter particles", "spaces between particles")),
    (8, 4, "Properties of Matter", "I", "MAT", "PROPS",
     "Utilizing physical properties of matter",
     _kw("physical properties", "density basic", "conductivity", "solubility",
         "hardness property", "using properties of matter")),
    (8, 5, "Sound", "I", "SOU", "INSTR",
     "Musical instruments and sound production",
     _kw("musical instrument", "vibrating membrane", "air column", "string instrument",
         "sound instrument", "music sound")),
    (8, 5, "Sound", "I", "SOU", "VIBRATE",
     "Vibration types that produce sound",
     _kw("vibration", "vibrate string", "vibrate air", "produce sound instrument",
         "pitch instrument")),
    (8, 6, "Magnets", "I", "MAG", "FIELD",
     "Magnetic poles, field, and compass",
     _kw("magnetic field", "compass", "geomagnetism", "poles of a magnet",
         "magnetic lines")),
    (8, 6, "Magnets", "I", "MAG", "TYPES",
     "Temporary and permanent magnets",
     _kw("temporary magnet", "permanent magnet", "electromagnet intro",
         "magnet types grade8")),
    (8, 7, "Measurements Associated with Electricity", "I", "ELE", "CURRENT",
     "Electric current and potential difference",
     _kw("electric current", "potential difference", "voltage", "ammeter",
         "voltmeter", "measurements electricity")),
    (8, 7, "Measurements Associated with Electricity", "I", "ELE", "RESIST",
     "Resistance of a conductor",
     _kw("resistance", "conductor resistance", "ohm", "resistor",
         "factors affecting resistance")),
    (8, 8, "Changes in Matter", "I", "CHA", "PHYSCHEM",
     "Physical and chemical changes",
     _kw("physical change", "chemical change", "change of state",
         "physical vs chemical")),
    (8, 8, "Changes in Matter", "I", "CHA", "COMBUST",
     "Combustion, tarnishing, and neutralisation",
     _kw("combustion", "burning", "tarnishing", "neutralisation", "rust",
         "chemical changes burning")),

    # ─── GRADE 8 Part II (9–15) ────────────────────────────────────────────
    (8, 9, "Human Organ Systems", "II", "HUM", "EXCRET",
     "Human excretory system",
     _kw("excretory", "excretion", "kidney", "urine", "urea", "sweat",
         "excretory products")),
    (8, 9, "Human Organ Systems", "II", "HUM", "NERVSKIN",
     "Nervous system and human skin",
     _kw("nervous system", "neuron", "brain", "spinal cord", "skin",
         "sense organ skin")),
    (8, 10, "Electricity", "II", "ELE", "CIRCUITS",
     "Simple circuits and cell/bulb connections",
     _kw("simple electric circuit", "series parallel grade8", "connecting cells",
         "connecting bulbs", "circuit components")),
    (8, 10, "Electricity", "II", "ELE", "EFFECTS",
     "Heating, light, magnetic, chemical effects of current",
     _kw("heating effect", "light effect current", "magnetic effect current",
         "chemical effect", "household electrical", "current controlling")),
    (8, 11, "Main Biological Processes in Plants", "II", "PHO", "PROCESS",
     "Photosynthesis process",
     _kw("photosynthesis", "chlorophyll", "glucose", "carbon dioxide plant",
         "oxygen plant", "raw materials photosynthesis")),
    (8, 11, "Main Biological Processes in Plants", "II", "PHO", "TRANSP",
     "Transportation, transpiration, and guttation",
     _kw("transportation plant", "xylem", "phloem", "transpiration",
         "guttation", "water transport plant")),
    (8, 12, "Life Cycles of Living Organisms", "II", "LIF", "ANIMAL",
     "Life cycles of animals",
     _kw("life cycle animal", "metamorphosis", "egg larva pupa",
         "life stages animal")),
    (8, 12, "Life Cycles of Living Organisms", "II", "LIF", "PLANT",
     "Life cycles of plants and their importance",
     _kw("life cycle plant", "seed germination", "pollination", "life stages plant",
         "importance of life cycles")),
    (8, 13, "Food Preservation", "II", "FOO", "METHODS",
     "Need and methods of food preservation",
     _kw("food preservation", "preserve food", "drying food", "refrigeration",
         "canning", "food preservative")),
    (8, 13, "Food Preservation", "II", "FOO", "LABEL",
     "Advantages of preservation and food labels",
     _kw("food label", "expiry date", "preservative advantages",
         "disadvantages of preservation")),
    (8, 14, "Solar System Phenomena", "II", "SOL", "SYSTEM",
     "The solar system and seasonal/lunar phenomena",
     _kw("solar system", "planet", "sun", "moon", "seasons", "phases of moon",
         "seasonal changes")),
    (8, 14, "Solar System Phenomena", "II", "SOL", "EXPLORE",
     "Exploring the universe, satellites, constellations",
     _kw("artificial satellite", "constellation", "universe", "space exploration",
         "astronomy basic")),
    (8, 15, "Natural Disasters", "II", "DIS", "HYDRO",
     "Drought, floods, and landslides",
     _kw("drought", "flood", "landslide", "earth slip", "natural disaster")),
    (8, 15, "Natural Disasters", "II", "DIS", "LIGHTNG",
     "Lightning and thundering",
     _kw("lightning", "thunder", "thunderstorm", "lightning safety grade8")),

    # ─── GRADE 9 Part I (1–9) ──────────────────────────────────────────────
    (9, 1, "Applications of Micro-organisms", "I", "MIC", "ENV",
     "Micro-organisms, substrates, and environments",
     _kw("micro-organisms", "microorganisms applications", "substrates",
         "environments of microbes")),
    (9, 1, "Applications of Micro-organisms", "I", "MIC", "EFFECTS",
     "Effects and applications of micro-organisms",
     _kw("useful microbes", "harmful microbes", "fermentation",
         "applications of microorganisms", "biotechnology basic")),
    (9, 2, "Eye and Ear", "I", "SEN", "EYE",
     "Structure and defects of the human eye",
     _kw("eye", "vision", "retina", "lens eye", "defects of vision",
         "myopia", "hypermetropia", "eye disease")),
    (9, 2, "Eye and Ear", "I", "SEN", "EAR",
     "Structure and defects of the human ear",
     _kw("ear", "hearing", "eardrum", "cochlea", "defects of ear",
         "auditory")),
    (9, 3, "Nature and Properties of Matter", "I", "MAT", "ELEMENTS",
     "Elements, compounds, and mixtures",
     _kw("element", "compound", "mixture", "chemical symbol",
         "properties of matter grade9", "nature of matter")),
    (9, 3, "Nature and Properties of Matter", "I", "MAT", "MIXTURES",
     "Mixtures and separation ideas",
     _kw("mixture", "homogeneous", "heterogeneous", "separate mixture",
         "solution mixture")),
    (9, 4, "Basic Concepts Associated with Force", "I", "FOR", "FORCE",
     "Force, magnitude, direction, and point of application",
     _kw("force", "magnitude of force", "direction of force", "point of application",
         "newton", "graphical force")),
    (9, 4, "Basic Concepts Associated with Force", "I", "FOR", "GRAPH",
     "Graphical representation of force",
     _kw("force diagram", "vector force", "arrow force", "represent force")),
    (9, 5, "Pressure Exerted by Solid", "I", "PRE", "PRESSURE",
     "Pressure and factors affecting pressure",
     _kw("pressure", "force area", "factors affecting pressure", "pascal",
         "unit of pressure")),
    (9, 5, "Pressure Exerted by Solid", "I", "PRE", "APPLY",
     "Changing pressure factors as needed",
     _kw("increase pressure", "decrease pressure", "sharp knife pressure",
         "applications of pressure")),
    (9, 6, "The Human Circulatory System", "I", "CIR", "HEART",
     "Structure of the heart; vessels",
     _kw("circulatory", "heart", "artery", "vein", "capillary", "blood vessel")),
    (9, 6, "The Human Circulatory System", "I", "CIR", "BLOOD",
     "Blood components and transfusion",
     _kw("blood", "red blood cell", "white blood cell", "plasma", "platelets",
         "blood transfusion", "blood group")),
    (9, 7, "Plant Growth Substances", "I", "PGS", "INTRO",
     "Introduction to plant growth substances",
     _kw("plant growth substance", "hormone plant", "auxin", "gibberellin",
         "cytokinin", "plant hormone")),
    (9, 7, "Plant Growth Substances", "I", "PGS", "ARTIF",
     "Uses of artificial growth substances",
     _kw("artificial growth substance", "weedicide", "rooting hormone",
         "plant growth regulator")),
    (9, 8, "Support and Movements of Organisms", "I", "MOV", "ANIMAL",
     "Bones, muscles, joints; animal movement",
     _kw("bone", "muscle", "joint", "skeleton", "movement animal",
         "support animals")),
    (9, 8, "Support and Movements of Organisms", "I", "MOV", "PLANT",
     "Support and movements of plants",
     _kw("tropism", "plant movement", "turgor", "support plants",
         "phototropism")),
    (9, 9, "The Evolutionary Process", "I", "EVO", "ORIGIN",
     "Origin of Earth and life",
     _kw("origin of earth", "origin of life", "evolution intro",
         "evolutionary process")),
    (9, 9, "The Evolutionary Process", "I", "EVO", "BIODIV",
     "Evolution and importance for biodiversity",
     _kw("evolution", "biodiversity evolution", "natural selection intro",
         "variation species")),

    # ─── GRADE 9 Part II (10–19) ───────────────────────────────────────────
    (9, 10, "Electrolysis", "II", "ELE", "LYSIS",
     "Electrolysis and solution changes by current",
     _kw("electrolysis", "electrolyte", "electrode", "anode", "cathode",
         "electric current solution")),
    (9, 10, "Electrolysis", "II", "ELE", "PLATE",
     "Electroplating",
     _kw("electroplating", "plating metal", "coat metal electrically")),
    (9, 11, "Density", "II", "DEN", "INTRO",
     "Introduction to density and units",
     _kw("density", "mass volume", "unit of density", "kg per m3",
         "relative density intro")),
    (9, 11, "Density", "II", "DEN", "HYDRO",
     "Hydrometers",
     _kw("hydrometer", "measure density liquid", "floating densitometer")),
    (9, 12, "Bio-diversity", "II", "BIO", "INTRO",
     "Bio-diversity and its importance",
     _kw("biodiversity", "bio-diversity", "species diversity",
         "importance of biodiversity")),
    (9, 12, "Bio-diversity", "II", "BIO", "ECO",
     "Ecosystems, threats, and built vs natural environments",
     _kw("ecosystem", "threats to biodiversity", "natural ecosystem",
         "built environment", "habitat loss")),
    (9, 13, "Artificial Environment and Green Concept", "II", "GRN", "CONCEPT",
     "Artificial environment and green concept",
     _kw("green concept", "artificial environment", "sustainable living",
         "environment friendly")),
    (9, 13, "Artificial Environment and Green Concept", "II", "GRN", "AGRI",
     "Agricultural and industrial processes (green)",
     _kw("agricultural process", "industrial process", "green agriculture",
         "eco friendly industry")),
    (9, 14, "Reflection and Refraction of Waves", "II", "WAV", "REFLECT",
     "Reflection of light and sound",
     _kw("reflection of light", "reflection of sound", "echo", "mirror reflection",
         "wave reflection")),
    (9, 14, "Reflection and Refraction of Waves", "II", "WAV", "REFRACT",
     "Refraction of light",
     _kw("refraction", "bending of light", "refractive index", "lens light",
         "critical angle", "total internal reflection")),
    (9, 15, "Simple Machines", "II", "MAC", "LEVER",
     "Lever and inclined plane",
     _kw("lever", "inclined plane", "simple machine", "mechanical advantage",
         "load effort")),
    (9, 15, "Simple Machines", "II", "MAC", "PULLEY",
     "Wheel and axle; pulleys",
     _kw("pulley", "wheel and axle", "movable pulley", "fixed pulley",
         "simple machines pulley")),
    (9, 16, "Nanotechnology and its Applications", "II", "NANO", "INTRO",
     "Nanometer and nanotechnology basics",
     _kw("nanotechnology", "nanometer", "nano scale", "nanoparticle")),
    (9, 16, "Nanotechnology and its Applications", "II", "NANO", "APPS",
     "Applications and future of nanotechnology",
     _kw("applications of nanotechnology", "nano medicine", "nano materials",
         "future nanotechnology")),
    (9, 17, "Lightning Accidents", "II", "LIG", "OCCUR",
     "How lightning occurs",
     _kw("lightning occurs", "thunder cloud", "static discharge lightning",
         "lightning formation")),
    (9, 17, "Lightning Accidents", "II", "LIG", "PREVENT",
     "Prevention of lightning accidents",
     _kw("lightning safety", "lightning conductor", "prevent lightning",
         "lightning accident")),
    (9, 18, "Natural Disasters", "II", "DIS", "TYPES",
     "Cyclones, earthquakes, tsunami, wild fires",
     _kw("cyclone", "earthquake", "tsunami", "wild fire", "natural disaster grade9")),
    (9, 18, "Natural Disasters", "II", "DIS", "WARMING",
     "Global warming and disasters",
     _kw("global warming", "climate disaster", "greenhouse",
         "relationship global warming disasters")),
    (9, 19, "Sustainable Use of Natural Resources", "II", "RES", "WATER",
     "Sustainable use of water",
     _kw("sustainable water", "water resources", "conserve water resources",
         "natural resources water")),
    (9, 19, "Sustainable Use of Natural Resources", "II", "RES", "MINERAL",
     "Sustainable use of minerals, rocks, and trees",
     _kw("sustainable minerals", "rocks resources", "trees conservation",
         "natural resources trees", "sustainable use")),
]


def topic_id(grade: int, chapter: int, domain: str, concept: str) -> str:
    return f"G{grade}_C{chapter}_{domain}_{concept}"


def build_rows() -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for grade, chapter, chapter_title, part, domain, concept, skill_label, keywords in CURRICULUM:
        tid = topic_id(grade, chapter, domain, concept)
        rows.append(
            {
                "grade": grade,
                "chapter": chapter,
                "chapter_title": chapter_title,
                "part": str(part),
                "topic_id": tid,
                "domain": domain,
                "concept": concept,
                "skill_label": skill_label,
                "curriculum_reference": f"Ch.{chapter}: {skill_label}",
                "keywords": keywords,
                "query_boost": f"Grade {grade} Chapter {chapter} ({chapter_title}). {skill_label}.",
            }
        )
    return rows


def write_excel(rows: list[dict[str, Any]], path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Skill Hierarchy"
    headers = [
        "Grade",
        "Chapter",
        "Chapter ID",
        "Chapter Title",
        "Part",
        "Core Concept",
        "Topic ID (Canonical)",
        "Curriculum Reference",
        "Domain",
        "Concept Code",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    last_grade = None
    for r in rows:
        if r["grade"] != last_grade:
            ws.append([f"GRADE {r['grade']}", "", "", "", "", "", "", "", "", ""])
            last_grade = r["grade"]
        ws.append(
            [
                r["grade"],
                r["chapter"],
                f"G{r['grade']}_C{r['chapter']}",
                r["chapter_title"],
                r["part"],
                r["chapter_title"],
                r["topic_id"],
                r["curriculum_reference"],
                r["domain"],
                r["concept"],
            ]
        )

    # Second sheet: chapter coverage summary
    ws2 = wb.create_sheet("Chapter Coverage")
    ws2.append(["Chapter ID", "Grade", "Chapter", "Chapter Title", "Part", "Skill Count", "Topic IDs"])
    for cell in ws2[1]:
        cell.font = Font(bold=True)
    from collections import defaultdict

    grouped: dict[tuple[int, int], list[dict[str, Any]]] = defaultdict(list)
    for r in rows:
        grouped[(r["grade"], r["chapter"])].append(r)
    for (g, c) in sorted(grouped.keys()):
        items = grouped[(g, c)]
        ws2.append(
            [
                f"G{g}_C{c}",
                g,
                c,
                items[0]["chapter_title"],
                items[0]["part"],
                len(items),
                ", ".join(x["topic_id"] for x in items),
            ]
        )

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    print(f"Wrote {path} ({len(rows)} skills)")


def write_chapter_csv(rows: list[dict[str, Any]], path: Path) -> None:
    """Shareable chapter_id catalog for Component 2 (one row per chapter)."""
    from collections import OrderedDict

    chapters: OrderedDict[str, dict[str, Any]] = OrderedDict()
    for r in rows:
        cid = f"G{int(r['grade'])}_C{int(r['chapter'])}"
        rec = chapters.setdefault(
            cid,
            {
                "chapter_id": cid,
                "grade": int(r["grade"]),
                "chapter": int(r["chapter"]),
                "chapter_title": r["chapter_title"],
                "topic_ids": [],
            },
        )
        rec["topic_ids"].append(r["topic_id"])

    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle)
        writer.writerow(
            [
                "chapter_id",
                "grade",
                "chapter",
                "chapter_title",
                "topic_id_1",
                "topic_id_2",
            ]
        )
        for rec in chapters.values():
            tids = list(rec["topic_ids"])
            while len(tids) < 2:
                tids.append("")
            writer.writerow(
                [
                    rec["chapter_id"],
                    rec["grade"],
                    rec["chapter"],
                    rec["chapter_title"],
                    tids[0],
                    tids[1],
                ]
            )
    print(f"Wrote {path} ({len(chapters)} chapters)")


def write_curriculum_py(rows: list[dict[str, Any]], path: Path) -> None:
    keywords = {r["topic_id"]: r["keywords"] for r in rows}
    boosts = {r["topic_id"]: r["query_boost"] for r in rows}
    meta = {
        r["topic_id"]: {
            "grade": r["grade"],
            "chapter": r["chapter"],
            "chapter_title": r["chapter_title"],
            "part": r["part"],
            "skill_label": r["skill_label"],
            "curriculum_reference": r["curriculum_reference"],
        }
        for r in rows
    }
    topic_ids = [r["topic_id"] for r in rows]

    content = f'''"""
Auto-generated curriculum catalog (do not hand-edit).

Regenerate with:
    python Scripts/build_full_curriculum.py

Format: G{{grade}}_C{{chapter}}_{{DOMAIN}}_{{CONCEPT}}
Skills: {len(topic_ids)} across grades 6–9 (full chapter coverage).
"""

from __future__ import annotations

import re
from typing import Any, Optional

# Canonical ordered topic IDs
TOPIC_IDS: list[str] = {json.dumps(topic_ids, indent=4)}

TOPIC_META: dict[str, dict[str, Any]] = {json.dumps(meta, indent=4)}

TOPIC_KEYWORDS: dict[str, list[str]] = {json.dumps(keywords, indent=4)}

TOPIC_QUERY_BOOST: dict[str, str] = {json.dumps(boosts, indent=4)}

FALLBACK_TOPIC_ID = {json.dumps(topic_ids[0] if topic_ids else "G6_C1_ORG_CHARS")}

OLD_TO_NEW_TOPIC_ID: dict[str, str] = {{
    # Best-effort migration from previous S-section IDs → chapter-aligned C IDs
    "G6_S1_ORG_CHARS": "G6_C1_ORG_CHARS",
    "G6_S1_ORG_CLASS": "G6_C1_ORG_DIFF",
    "G6_S2_MAT_STATES": "G6_C2_MAT_STATES",
    "G6_S2_MAT_PROPS": "G6_C2_MAT_PROPS",
    "G6_S4_ENE_SOURCES": "G6_C4_ENE_SOURCES",
    "G6_S8_ELE_CIRCUITS": "G6_C8_ELE_CIRCUITS",
    "G6_S8_ELE_CONDINS": "G6_C8_ELE_CONDINS",
    "G7_S1_PLA_DIVER": "G7_C1_PLA_DIVER",
    "G7_S1_PLA_CLASSIF": "G7_C1_PLA_CLASSIF",
    "G7_S2_STA_CHARGES": "G7_C2_STA_CHARGES",
    "G7_S2_STA_CAPACIT": "G7_C2_STA_CAPACIT",
    "G7_S3_ELE_SOURCES": "G7_C3_ELE_SOURCES",
    "G7_S3_ELE_CURRENTS": "G7_C3_ELE_CURRENTS",
    "G7_S4_WAT_SOLVENT": "G7_C4_WAT_SOLVENT",
    "G7_S4_WAT_COOLANT": "G7_C4_WAT_COOLANT",
    "G7_S5_ACI_IDENTIF": "G7_C5_ACI_IDENTIF",
    "G7_S5_ACI_INDICAT": "G7_C5_ACI_INDICAT",
    "G7_S6_ANI_CLASSIF": "G7_C6_ANI_CLASSIF",
    "G7_S6_ANI_ADAPTAT": "G7_C6_ANI_ADAPTAT",
    "G7_S7_ENE_FORMS": "G7_C7_ENE_FORMS",
    "G7_S7_ENE_TRANSF": "G7_C7_ENE_TRANSF",
    "G7_S8_EAR_STRUCT": "G7_C8_EAR_STRUCT",
    "G7_S8_EAR_TECTON": "G7_C8_EAR_TECTON",
    "G7_S9_LIG_SHADOWS": "G7_C9_LIG_SHADOWS",
    "G7_S9_LIG_MIRRORS": "G7_C9_LIG_MIRRORS",
    "G7_S10_MIC_LIGHT": "G7_C10_MIC_LIGHT",
    "G7_S10_MIC_ELECTR": "G7_C10_MIC_ELECTR",
    "G8_S1_BIO_DIVER": "G8_C1_MIC_INTRO",
    "G8_S1_BIO_CLASSIF": "G8_C2_ANI_VERT",
    "G8_S2_TIS_PLANT": "G8_C3_PLA_LEAVES",
    "G8_S2_TIS_ANIMAL": "G8_C9_HUM_NERVSKIN",
    "G8_S3_PHO_PROCESS": "G8_C11_PHO_PROCESS",
    "G8_S3_PHO_IMPORT": "G8_C11_PHO_TRANSP",
    "G8_S4_MAT_ELEMENTS": "G8_C4_MAT_PARTICLE",
    "G8_S4_MAT_COMPOUNDS": "G9_C3_MAT_ELEMENTS",
    "G8_S5_MAT_DENSITY": "G9_C11_DEN_INTRO",
    "G8_S5_MAT_THERMAL": "G8_C4_MAT_PROPS",
    "G8_S6_CHA_PHYSICAL": "G8_C8_CHA_PHYSCHEM",
    "G8_S6_CHA_BURNING": "G8_C8_CHA_COMBUST",
    "G8_S7_FOR_TYPES": "G9_C4_FOR_FORCE",
    "G8_S7_FOR_PRESSURE": "G9_C5_PRE_PRESSURE",
    "G8_S8_STA_PHENOM": "G7_C2_STA_CHARGES",
    "G8_S8_STA_LIGHTNG": "G8_C15_DIS_LIGHTNG",
    "G9_S1_SYS_DIGEST": "G7_C12_BIO_SYSTEMS",
    "G9_S1_SYS_CIRCUL": "G9_C6_CIR_HEART",
    "G9_S2_RHY_EARTH": "G8_C14_SOL_SYSTEM",
    "G9_S2_RHY_CLIMATE": "G6_C11_WEA_WEATHER",
    "G9_S3_LIG_REFRAC": "G9_C14_WAV_REFRACT",
    "G9_S3_LIG_LENSES": "G9_C14_WAV_REFRACT",
    "G9_S4_SOU_PROPAG": "G7_C11_SOU_PROPAG",
    "G9_S4_SOU_HEARING": "G9_C2_SEN_EAR",
    "G9_S5_HEA_EXPANS": "G7_C14_HEA_MEASURE",
    "G9_S5_HEA_TRANSF": "G7_C14_HEA_TRANSF",
    "G9_S6_NAT_ATOMS": "G9_C3_MAT_ELEMENTS",
    "G9_S6_NAT_CONFIG": "G9_C3_MAT_ELEMENTS",
    "G9_S7_ACI_SALTS": "G7_C5_ACI_IDENTIF",
    "G9_S7_ACI_NEUTRAL": "G7_C5_ACI_INDICAT",
}}


def normalize_topic_id(topic_id: str) -> str:
    """Map legacy S-IDs to current C-IDs when possible."""
    tid = str(topic_id or "").strip()
    return OLD_TO_NEW_TOPIC_ID.get(tid, tid)


_CHAPTER_ID_RE = re.compile(r"^G(\\d+)[_-]C(\\d+)$", re.IGNORECASE)
_TOPIC_CHAPTER_PREFIX_RE = re.compile(r"^G(\\d+)_C(\\d+)(?:_|$)", re.IGNORECASE)


def canonical_chapter_id(grade: int, chapter: int) -> str:
    """Return the shared chapter key, e.g. G6_C8."""
    return f"G{{int(grade)}}_C{{int(chapter)}}"


def chapter_id_for_topic(topic_id: str) -> Optional[str]:
    """Map a canonical (or legacy) topic_id to G{{grade}}_C{{chapter}}."""
    tid = normalize_topic_id(topic_id)
    meta = TOPIC_META.get(tid)
    if meta:
        return canonical_chapter_id(meta["grade"], meta["chapter"])
    match = _TOPIC_CHAPTER_PREFIX_RE.match(tid)
    if match:
        return canonical_chapter_id(int(match.group(1)), int(match.group(2)))
    return None


def normalize_chapter_id(value: str) -> Optional[str]:
    """
    Normalize a chapter key to G{{grade}}_C{{chapter}}.

    Accepts ``G6_C8``, ``g6-c8``, or a full topic_id such as ``G6_C8_ELE_CIRCUITS``.
    Returns None when the string cannot be parsed.
    """
    raw = str(value or "").strip()
    if not raw:
        return None
    tid = normalize_topic_id(raw)
    if tid in TOPIC_META:
        return chapter_id_for_topic(tid)
    match = _CHAPTER_ID_RE.fullmatch(raw)
    if match:
        return canonical_chapter_id(int(match.group(1)), int(match.group(2)))
    return chapter_id_for_topic(raw)


def topic_ids_for_chapter(chapter_id: str) -> list[str]:
    """Return ordered canonical topic_ids for one chapter (usually two skills)."""
    cid = normalize_chapter_id(chapter_id)
    if not cid:
        return []
    match = _CHAPTER_ID_RE.fullmatch(cid)
    if not match:
        return []
    grade = int(match.group(1))
    chapter = int(match.group(2))
    return [
        tid
        for tid in TOPIC_IDS
        if int(TOPIC_META[tid]["grade"]) == grade and int(TOPIC_META[tid]["chapter"]) == chapter
    ]


def resolve_chapter_scope(chapter_ids: list[str]) -> dict[str, Any]:
    """
    Deduplicate and resolve chapter keys to topic_ids.

    Unknown / empty chapters are listed in ``unknown_chapter_ids`` and skipped.
    """
    resolved: list[str] = []
    unknown: list[str] = []
    topics_by_chapter: dict[str, list[str]] = {{}}
    all_topics: list[str] = []
    seen: set[str] = set()
    for raw in chapter_ids:
        cid = normalize_chapter_id(raw)
        topics = topic_ids_for_chapter(cid) if cid else []
        if not cid or not topics:
            unknown.append(str(raw))
            continue
        if cid in seen:
            continue
        seen.add(cid)
        resolved.append(cid)
        topics_by_chapter[cid] = topics
        all_topics.extend(topics)
    return {{
        "chapter_ids": resolved,
        "unknown_chapter_ids": unknown,
        "topic_ids": all_topics,
        "topics_by_chapter": topics_by_chapter,
    }}


def chapters_covered() -> dict[int, list[int]]:
    out: dict[int, list[int]] = {{}}
    for tid, m in TOPIC_META.items():
        out.setdefault(int(m["grade"]), [])
        ch = int(m["chapter"])
        if ch not in out[int(m["grade"])]:
            out[int(m["grade"])].append(ch)
    for g in out:
        out[g] = sorted(out[g])
    return out
'''
    path.write_text(content, encoding="utf-8")
    print(f"Wrote {path} ({len(topic_ids)} topic_ids)")


def validate(rows: list[dict[str, Any]]) -> None:
    expected = {
        6: list(range(1, 12)),
        7: list(range(1, 20)),
        8: list(range(1, 16)),
        9: list(range(1, 20)),
    }
    from collections import defaultdict

    found: dict[int, set[int]] = defaultdict(set)
    ids: set[str] = set()
    for r in rows:
        found[r["grade"]].add(r["chapter"])
        assert r["topic_id"] not in ids, f"duplicate {r['topic_id']}"
        ids.add(r["topic_id"])
        assert re.match(r"^G[6-9]_C\d+_[A-Z]+_[A-Z0-9]+$", r["topic_id"]), r["topic_id"]

    for g, chapters in expected.items():
        missing = [c for c in chapters if c not in found[g]]
        extra = sorted(found[g] - set(chapters))
        if missing:
            raise SystemExit(f"Grade {g} missing chapters: {missing}")
        if extra:
            raise SystemExit(f"Grade {g} unexpected chapters: {extra}")
    print(
        f"Validation OK: {len(rows)} skills; "
        f"G6 chapters {sorted(found[6])}; "
        f"G7 {sorted(found[7])}; "
        f"G8 {sorted(found[8])}; "
        f"G9 {sorted(found[9])}"
    )


def main() -> None:
    rows = build_rows()
    validate(rows)
    primary = DATA_DIR / "Skill-Heirarchies-G6-G9.xlsx"
    shareable = DATA_DIR / "Skill-Heirarchies-G6-G9-Full-Chapters.xlsx"
    write_excel(rows, shareable)
    try:
        write_excel(rows, primary)
    except PermissionError:
        alt = DATA_DIR / "Skill-Heirarchies-G6-G9-UPDATED.xlsx"
        write_excel(rows, alt)
        print(
            f"WARNING: could not overwrite {primary.name} (file may be open). "
            f"Wrote {alt.name} instead — close Excel and rename/copy over when ready."
        )
    write_curriculum_py(rows, BACKEND_DIR / "curriculum_topics.py")
    write_chapter_csv(rows, DATA_DIR / "chapter_ids_g6_g9.csv")
    print("Done.")


if __name__ == "__main__":
    main()
